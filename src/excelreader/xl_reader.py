import openpyxl
from openpyxl.workbook.workbook import Workbook

work_book=openpyxl.load_workbook("C:\\Users\\mages\\PycharmProjects\\PythonSeleniumBasics\\test_data.xlsx") #loading the xl sheet using openpyxl
sheet=work_book["test"] #going to the specific sheet

print(sheet.cell(1,1).value) #gets the cell value of row 1 column 1

#another way of printing the cell value
print(sheet['A1'].value) #this is not a recommend approach since we don't know the column name

#looping through column1
print("Column A")
for cell in sheet['A']:
    print(cell.value)

#getting total rows
print(sheet.max_row)

#getting total columns
print(sheet.max_column)

#getting all the values from xl

for row in range(0,sheet.max_row):
    row_count=row+1
    print(row_count)
    for cell in sheet[row_count]:
        print(cell," ",cell.value)

#another way of iterating over a sheet

for row in sheet.iter_rows(1,sheet.max_row): #iter_rows will iterate over the last row
    for cell in row:
        print(cell," ",cell.value)

#If we want to store the data from the xl for our use case, we have to store them in a list of tuples
print("Storing data in a list of tuple")
sheet_cells=[] #created a empty list that's going to hold a list of tuples

for row in sheet.iter_rows():
    row_cells=[] #created a list to hold individual row elements
    for cell in row:
        row_cells.append(cell.value) #adding cell element from each row
    sheet_cells.append(tuple(row_cells)) #adding the rows of values into a list after converting them into a tuple

print(sheet_cells)

#extracting the data from the list of tuples
print("Extracting the data from the list of tuples: Ignore the headers ")
for i in range(1,len(sheet_cells)): #index starts from 0 in tuple
    print(sheet_cells[i])
    print(sheet_cells[i][0]," | ",sheet_cells[i][1]," | ",sheet_cells[i][2])

# #creating a sheet
# work_book.create_sheet("datasheet")
sheet2=work_book['datasheet']
#
# #adding data to the empty sheet
# sheet2.cell(1,1).value="Tester"
# print(sheet2.cell(1,1).value)
#
#work_book.save("C:\\Users\\mages\\PycharmProjects\\PythonSeleniumBasics\\test_data.xlsx") #saving the workbook

#writing the list of tuples to the sheet
testdata=[('username', 'password', 'status'), ('test_user1', 'Tester@123', 'pass'), ('test_user2', 'Tester@111', 'fail')]
for item in testdata:
    sheet2.append(item)

work_book.save("C:\\Users\\mages\\PycharmProjects\\PythonSeleniumBasics\\test_data.xlsx") #saving the workbook
